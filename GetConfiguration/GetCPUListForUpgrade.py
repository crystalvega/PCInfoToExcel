from openpyxl import load_workbook

def format_cpu(cpu):
    wordsdeleteintel = ['12th Gen ','11th Gen ','10th Gen ', '(R)', '(TM)', "CPU "]
    wordsdeleteamd = [' with Radeon Vega Graphics', " Mobile", " Quad-Core Processor           "]
    for word in wordsdeleteintel:
        if word in cpu:
            cpu = cpu.replace(word, "")
            if " @" in cpu:
                cpu = cpu.split(' @')[0]
    for word in wordsdeleteamd:
        if word in cpu:
            cpu = cpu.replace(word, "")
    return cpu


def get_cpu_info(cpu):
    cpu_info = ['']*6
    wb = load_workbook('C:/Windows/Temp/dataofpack1/cfg/db.xlsx')
    sheet = wb['CPU']
    cpu = format_cpu(cpu)
    i = 1
    ipspec = 0
    while i < 3519:
        if sheet.cell(row = i, column = 1).value in cpu:
            cpu_info[5] = i
            while ipspec < 5:
                cpu_info[ipspec] = sheet.cell(row = i, column = ipspec+1).value
                ipspec += 1
        i += 1
    return cpu_info

def get_cpu_upgrade(procinf):
    wb = load_workbook('C:/Windows/Temp/dataofpack1/cfg/db.xlsx')
    sheet = wb['CPU']
    i = 3519
    allcpus = 0
    lastcpu = ""
    cpusret = []
    while i != 1:
        if sheet.cell(row = i, column = 4).value == procinf[3] and sheet.cell(row = i, column = 2).value > procinf[1]:
            if allcpus == 0:
                allcpus = sheet.cell(row = i, column = 1).value
            else:
                lastcpu = sheet.cell(row = i, column = 1).value
                allcpus = allcpus + ', ' + lastcpu
        i -= 1
    cpusret = [allcpus, lastcpu]
    return cpusret