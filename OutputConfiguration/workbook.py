from pathlib import Path
import xlsxwriter as xlwr
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

def create(allconfiguration, diskconfiguration):
    fileName = "tablica.xlsx"
    fileObj = Path(fileName)
    raw = 0

    if fileObj.is_file() == False:
        workbook = xlwr.Workbook('tablica.xlsx', {'strings_to_numbers': True})
        worksheet = workbook.add_worksheet('Общие характеристики')
        worksheet2 = workbook.add_worksheet('Характеристики дисков')
        bold = workbook.add_format({'bold': True})
        while raw < len(allconfiguration):
            worksheet.write(0, raw, allconfiguration[raw][0], bold)
            raw += 1
        raw = 0
        while raw < 5:
            if raw < 4:
                worksheet2.write(0, raw, diskconfiguration[raw][0], bold)
            else:
                i = 0
                while i<len(diskconfiguration[4]):
                    worksheet2.write(0, raw+i, diskconfiguration[raw][i][0], bold)
                    i+=1
            raw += 1
        workbook.close()
        
def smart_cursor(allconfiguration):
    wb = load_workbook('./tablica.xlsx')
    sheet = wb['Общие характеристики']
    check = 'V'
    rowend = 1
    column = 0
    columnstart = 0
    columnend = 0
    
    while check == 'V':
        rowend += 1
        cell_obj = sheet.cell(row=rowend, column=1)
        type(cell_obj)
        check = cell_obj.value
    rowend -=1
    
    while column < len(allconfiguration):
        column +=1 
        cell_obj = sheet.cell(row=1, column=column)
        type(cell_obj)
        if cell_obj.value != allconfiguration[column-1][0] and cell_obj.value == "Операционная система":
            columnend = column
            columnstart = column
            while sheet.cell(row=1,column=columnend).value != None:
                columnend += 1
            while sheet.cell(row=1,column=column).value != allconfiguration[columnstart-1][0]:
                columnstart += 1
            for i in range(1, rowend):
                for j in range(column, columnend):
                    sheet.cell(row=i,column=j+columnstart).value = sheet.cell(row=i,column=j).value
    wb.save(filename = 'tablica.xlsx')
    
        
def configuration(allconfiguration, diskconfiguration):
    wb = load_workbook('./tablica.xlsx')
    sheet = wb['Общие характеристики']

    detected = 0
    row = 1
    check = 'V'
    column = 1
    hyperlink0 = 0
    hyperlink1 = 0
    hyperlink2 = 0


    while check == 'V':
        row += 1
        cell_obj = sheet.cell(row=row, column=1)
        type(cell_obj)
        check = cell_obj.value
        
        
    while column < len(allconfiguration):
        sheet.cell(row = row, column = column).value = allconfiguration
        raw += 1
    detected = 1
"""    sheet.cell(row = rownumb, column = 25).value = allcpus
    sheet.cell(row = rownumb, column = 26).value = " "
    hyperlink0 = rownumb"""